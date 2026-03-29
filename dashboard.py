"""
dashboard.py — Dashboard blueprint
Handles: main page, CV upload/analysis, usage API
"""

import os
import uuid
import shutil
import json
import io
from pathlib import Path
from datetime import datetime

from flask import (Blueprint, render_template, request, jsonify,
                   Response, send_file, stream_with_context, current_app)
from flask_login import login_required, current_user

from models import db, User
from analyzer import analyze_all_stream

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

dash_bp = Blueprint("dashboard", __name__)

UPLOAD_FOLDER = Path("uploads")
ALLOWED_EXTS  = {".pdf", ".docx", ".doc", ".txt"}
UPLOAD_FOLDER.mkdir(exist_ok=True)

# In-memory session results
_sessions: dict[str, list] = {}


# ─── MAIN DASHBOARD ───────────────────────────────────────────────────────────

@dash_bp.route("/")
@login_required
def index():
    current_user.reset_daily_usage_if_needed()
    current_user.check_and_expire_trial()
    return render_template("dashboard/index.html", user=current_user)


@dash_bp.route("/upgrade")
@login_required
def upgrade():
    paddle_vendor_id  = current_app.config["PADDLE_VENDOR_ID"]
    paddle_product_id = current_app.config["PADDLE_PRODUCT_ID"]
    return render_template(
        "dashboard/upgrade.html",
        user=current_user,
        paddle_vendor_id=paddle_vendor_id,
        paddle_product_id=paddle_product_id,
    )


# ─── API: USAGE STATUS ────────────────────────────────────────────────────────

@dash_bp.route("/api/usage")
@login_required
def api_usage():
    current_user.reset_daily_usage_if_needed()
    current_user.check_and_expire_trial()
    return jsonify({
        "plan":          current_user.effective_plan,
        "badge":         current_user.plan_badge,
        "daily_limit":   current_user.daily_limit,
        "used":          current_user.daily_usage_count,
        "remaining":     current_user.usage_remaining,
        "percent":       current_user.usage_percent,
        "trial_active":  current_user.is_trial_active,
        "trial_days_left": current_user.trial_days_left,
    })


# ─── UPLOAD ───────────────────────────────────────────────────────────────────

@dash_bp.route("/upload", methods=["POST"])
@login_required
def upload():
    sid    = request.form.get("session_id") or str(uuid.uuid4())
    folder = UPLOAD_FOLDER / str(current_user.id) / sid
    folder.mkdir(parents=True, exist_ok=True)

    saved = []
    for f in request.files.getlist("cvs"):
        if Path(f.filename).suffix.lower() in ALLOWED_EXTS:
            dest = folder / f.filename
            f.save(dest)
            saved.append({"name": Path(f.filename).stem, "filename": f.filename})

    return jsonify({"session_id": sid, "uploaded": saved, "count": len(saved)})


# ─── ANALYZE — SSE streaming ──────────────────────────────────────────────────

@dash_bp.route("/analyze")
@login_required
def analyze():
    current_user.reset_daily_usage_if_needed()
    current_user.check_and_expire_trial()

    api_key   = current_app.config["RESUMEPARSER_API_KEY"]
    sid       = request.args.get("session_id", "")
    job_title = request.args.get("job_title", "")
    job_desc  = request.args.get("job_desc", "")

    if not api_key:
        return jsonify({"error": "Clé API non configurée."}), 503

    folder   = UPLOAD_FOLDER / str(current_user.id) / sid
    cv_paths = [str(p) for p in folder.iterdir() if p.suffix.lower() in ALLOWED_EXTS] if folder.exists() else []

    if not cv_paths:
        return jsonify({"error": "Aucun CV trouvé."}), 400

    # Check limit BEFORE starting
    cv_count = len(cv_paths)
    if not current_user.can_analyze(cv_count):
        remaining = current_user.usage_remaining
        return jsonify({
            "error": f"Limite atteinte. Il vous reste {remaining} analyse(s) aujourd'hui.",
            "limit_reached": True,
            "remaining": remaining,
            "plan": current_user.effective_plan,
        }), 429

    user_id = current_user.id
    _sessions[sid] = []

    def generate():
        yield f"data: {json.dumps({'type':'start','total':len(cv_paths)})}\n\n"

        analyzed = 0
        for done, total, result in analyze_all_stream(cv_paths, api_key, job_title, job_desc):
            _sessions[sid].append(result)

            if result.get("_statut") == "OK":
                analyzed += 1
                # Increment usage per CV in DB
                user = User.query.get(user_id)
                if user:
                    user.increment_usage(1)

            safe = {k: v for k, v in result.items() if k not in ("_balance",)}
            yield f"data: {json.dumps({'type':'result','done':done,'total':total,'result':safe}, ensure_ascii=False)}\n\n"

        ok  = [r for r in _sessions[sid] if r.get("_statut") == "OK"]
        avg = round(sum(r.get("score_global", 0) for r in ok) / len(ok), 1) if ok else 0

        # Fetch fresh usage from DB
        user = User.query.get(user_id)
        yield f"data: {json.dumps({'type':'done','ok':len(ok),'total':len(cv_paths),'avg':avg,'used':user.daily_usage_count,'limit':user.daily_limit})}\n\n"

        _cleanup(user_id, sid)

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ─── EXPORT ───────────────────────────────────────────────────────────────────

@dash_bp.route("/export/excel/<sid>")
@login_required
def export_excel(sid):
    results   = _sessions.get(sid, [])
    ok        = sorted([r for r in results if r.get("_statut") == "OK"],
                       key=lambda x: x.get("score_global", 0), reverse=True)
    job_title = request.args.get("job_title", "Poste")

    wb  = _build_excel(ok, job_title)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"RecrutAI_{job_title.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@dash_bp.route("/export/csv/<sid>")
@login_required
def export_csv(sid):
    results = _sessions.get(sid, [])
    ok      = sorted([r for r in results if r.get("_statut") == "OK"],
                     key=lambda x: x.get("score_global", 0), reverse=True)
    lines   = ["Rang,Nom,Score,Recommandation,Expérience,Formation,Email"]
    for i, r in enumerate(ok, 1):
        row = [str(i), r.get("nom",""), str(r.get("score_global","")),
               r.get("recommandation",""), str(r.get("experience_annees","")),
               r.get("formation",""), r.get("email","")]
        lines.append(",".join(f'"{v}"' for v in row))
    buf = io.BytesIO(("\ufeff" + "\n".join(lines)).encode("utf-8"))
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="RecrutAI_resultats.csv",
                     mimetype="text/csv")


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def _cleanup(user_id, sid):
    folder = UPLOAD_FOLDER / str(user_id) / sid
    if folder.exists():
        shutil.rmtree(folder, ignore_errors=True)


def _b():
    t = Side(style="thin", color="D4C4AE")
    return Border(left=t, right=t, top=t, bottom=t)

def _h(ws, row, col, val, bg="3D2B1A", fc="FAF7F2", bold=True, align="center"):
    c = ws.cell(row=row, column=col)
    c.value = val; c.font = Font(name="Calibri", bold=bold, size=10, color=fc)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center"); c.border = _b()

def _d(ws, row, col, val, fc="1C1208", bg="FFFDF9", bold=False, align="left"):
    c = ws.cell(row=row, column=col)
    c.value = val; c.font = Font(name="Calibri", size=10, color=fc, bold=bold)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center"); c.border = _b()

def _sc(s):
    try:
        v = float(s)
        return ("D8F3DC","2D6A4F") if v>=8 else ("FFF3CD","856404") if v>=6 else ("FDE8E8","7D2226")
    except: return ("F3EDE3","6B5040")

def _rc(r):
    return ("D8F3DC","2D6A4F") if "recommandé" in str(r).lower() else \
           ("FFF3CD","856404") if "considérer" in str(r).lower() else ("FDE8E8","7D2226")

def _build_excel(results, job_title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "📊 Résultats"
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([2,5,26,18,12,10,10,10,22,2],1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("B1:J1"); ws.row_dimensions[1].height = 42
    c = ws["B1"]
    c.value = f"🎯  RecrutAI SaaS — {job_title}  |  {datetime.now().strftime('%d/%m/%Y')}"
    c.font = Font(name="Calibri", bold=True, size=15, color="B07D4E")
    c.fill = PatternFill("solid", fgColor="1C1208")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[3].height = 20
    for ci, h in enumerate(["#","Candidat","Poste","Exp.","Score","Adéq.","Niveau","Recommandation","Résumé"],2):
        _h(ws, 3, ci, h)

    medals = {1:"🥇",2:"🥈",3:"🥉"}
    for ri, r in enumerate(results, 4):
        ws.row_dimensions[ri].height = 30
        bg_r = "FAF7F2" if ri%2==0 else "FFFDF9"
        score = r.get("score_global",0); sbg,sfc = _sc(score); rbg,rfc = _rc(r.get("recommandation",""))
        rank = ri-3
        _d(ws,ri,2, medals.get(rank,rank), fc="B07D4E" if rank<=3 else "A08878", bg=bg_r, bold=rank<=3, align="center")
        _d(ws,ri,3, r.get("nom",""), fc="1C1208", bg=bg_r, bold=rank<=3)
        _d(ws,ri,4, r.get("poste_actuel",""), fc="5C4030", bg=bg_r)
        _d(ws,ri,5, f"{r.get('experience_annees','?')} ans", fc="5C4030", bg=bg_r, align="center")
        _d(ws,ri,6, f"{score}/10", fc=sfc, bg=sbg, bold=True, align="center")
        _d(ws,ri,7, f"{r.get('adequation_poste',0)}%", fc="5C4030", bg=bg_r, align="center")
        _d(ws,ri,8, r.get("niveau",""), fc="5C4030", bg=bg_r, align="center")
        _d(ws,ri,9, r.get("recommandation",""), fc=rfc, bg=rbg, bold=True, align="center")
        _d(ws,ri,10, str(r.get("resume_recruteur",""))[:90], fc="6B5040", bg=bg_r)
    ws.freeze_panes = "B4"
    return wb
